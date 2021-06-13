# -*- coding:utf-8 -*-
import os
import re
import time
import zipfile
from openpyxl import load_workbook
import pandas as pd
from settings import Config
from flask import request, send_from_directory
from . import ac_mod

# 文件保存路径
initial_data_path = Config.ROOT_PATH + './data/ACM/initial_data'
deal_data_path = Config.ROOT_PATH + './data/ACM/result_data/deal_data'
zip_data_path = Config.ROOT_PATH + './data/ACM/result_data/zip_data'
data_list_path = Config.ROOT_PATH + './data/ACM/data_list'
template_data_list_path = data_list_path + './template'
initial_data_template_file = initial_data_path + './其他.xlsx'


# —————————————————————————————————————————————————批量生成模型————————————————————————————————
def batch_create_mod(part_name, part_number, system_name, part_type):
    if part_type == 'nan':
        part_type = '其他'
    read_file = initial_data_path+'/'+part_type+'.xlsx'
    work_book = load_workbook(read_file)
    sheet = ['Model', 'Model_Y', 'Model_N', '01Assembly', '04Corrective Action-Cost']
    for i in sheet:
        if i == '01Assembly':
            sh = work_book[i]
            col = 1
            while True:
                if col == 6:
                    break
                elif sh.cell(2, col).value == '[Part_No]':
                    sh.cell(2, col).value = part_number
                elif sh.cell(2, col).value == '[Part_Name]':
                    sh.cell(2, col).value = part_name
                elif sh.cell(2, col).value == '[Sys_Name]':
                    sh.cell(2, col).value = system_name
                col += 1
        elif i == '04Corrective Action-Cost':
            sh = work_book[i]
            row = 2
            while True:
                if re.match(r'.*\[Part_Name\].*', sh.cell(row, 1).value):
                    s = sh.cell(row, 1).value.replace('[Part_Name]', part_name)
                    sh.cell(row, 1).value = s
                if row == sh.max_row:
                    break
                row += 1
        else:
            sh = work_book[i]
            col = 4
            while True:
                if sh.cell(1, col).value == '输出':
                    break
                elif re.match(r'.*\[Part_Name\].*', sh.cell(1, col).value):
                    s = sh.cell(1, col).value.replace('[Part_Name]', part_name)
                    sh.cell(1, col).value = s
                col += 1
    save_file_name = str(part_number)+'_'+part_name+'_'+time.strftime("%Y%m%d%H%M%S", time.localtime())[:8]+'_Draft.xlsx'
    work_book.save(deal_data_path + '/' + save_file_name)
    return save_file_name


# —————————————————————————————————————————————————模板文件下载————————————————————————————————
@ac_mod.route("/batch_build_mod/batch_download_template", methods=['GET'])
def batch_download_template():
    filename = '批量建模清单模板.xlsx'
    directory = template_data_list_path
    return send_from_directory(directory, filename, as_attachment=True)


# ———————————————————————————————————————————————上传模型清单接口————————————————————————————————
@ac_mod.route('/batch_build_mod/batch_upload_file', methods=['POST'])
def batch_upload_file():
    file = request.files['file']
    file_name = file.filename
    data_save = load_workbook(file)
    data = pd.read_excel(file)
    for i in data.columns.values.tolist():
        if i not in ['序号', '部件名称', '部件编号', '部件类型', '所属系统']:
            return {"code": 0, "state": "error", "msg": "部件清单有误，请修改后重新上传！"}
    data = data[['部件编号', '部件名称', '部件类型', '所属系统']]
    data.dropna(axis=0, how='all', inplace=True)
    data = data.astype({'部件类型': 'str'})
    file_name_list = os.listdir(initial_data_path)
    new_file_name_list = [i.rsplit('.', 1)[0] for i in file_name_list]
    part_type_list = data['部件类型'].values.tolist()
    for i in part_type_list:
        if i not in new_file_name_list:
            if i != 'nan':
                return {"code": 0, "state": "error", "msg": "部件类型输入错误！"}
    data1 = data[['部件编号', '部件名称', '所属系统']]
    if len(data1[data1.isnull().T.any()]) > 0:
        return {"code": 0, "state": "error", "msg": "部件清单有误，请修改后重新上传！"}
    data_save.save(os.path.join(data_list_path, file_name))
    return {"code": 1, "state": "success", "msg": "上传成功！"}


# —————————————————————————————————————————————————生成模型接口————————————————————————————————
@ac_mod.route('/batch_build_mod/create_mod', methods=['GET'])
def create_mod():
    file_name = request.args.get('fileName')
    file_path = data_list_path + "\\" + file_name
    data = pd.read_excel(file_path, sheet_name=0, header=0, index_col=0)
    data = data[['部件编号', '部件名称', '部件类型', '所属系统']]
    data.dropna(axis=0, how='all', inplace=True)
    data = data.astype({'部件编号': 'int', '部件类型': 'str'})
    save_file_name_list = []
    for i in range(data.shape[0]):
        save_file_name = batch_create_mod(part_name=data.iloc[i, 1], part_number=data.iloc[i, 0], system_name=data.iloc[i, 3], part_type=data.iloc[i, 2])
        save_file_name_list.append(save_file_name)
    zip_name = file_name.rsplit('.', 1)[0] + ".zip"
    zip_file = zipfile.ZipFile(zip_data_path + "/" + zip_name, "w")
    for save_file_name in save_file_name_list:
        zip_file.write(deal_data_path + "/" + save_file_name, save_file_name, zipfile.ZIP_DEFLATED)
    zip_file.close()
    return send_from_directory(zip_data_path, zip_name, as_attachment=True)

