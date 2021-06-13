# -*- coding:utf-8 -*-
import os
import re
import time
from openpyxl import load_workbook
from settings import Config
from flask import request, send_from_directory
from . import ac_mod

# 文件保存路径
initial_data_path = Config.ROOT_PATH + './data/ACM/initial_data'
deal_data_path = Config.ROOT_PATH + './data/ACM/result_data/deal_data'


# —————————————————————————————————————————————————发送模型类型————————————————————————————————
@ac_mod.route('/single_build_mod/send_file_name', methods=['GET'])
def send_file_name():
    file_name_list = os.listdir(initial_data_path)
    last_file_name_list = []
    for i in file_name_list:
        if not re.match(r'~\$.*', i):
            last_file_name_list.append(i.rsplit('.', 1)[0])
    options = [{'value': x} for x in last_file_name_list]
    data = {"code": 1, "state": "success", "options": options}
    return data


# —————————————————————————————————————————————————生成模型接口————————————————————————————————
@ac_mod.route('/single_build_mod/single_create_mod', methods=['GET'])
def single_create_mod():
    part_name = request.args.get('partName')
    part_number = request.args.get('partNumber')
    system_name = request.args.get('systemName')
    part_type = request.args.get('partType')
    if part_type == '':
        part_type = '其他'
    read_path = initial_data_path+'\\'+part_type+'.xlsx'
    work_book = load_workbook(read_path)
    sheet = ['Model', 'Model_Y', 'Model_N', '01Assembly', '04Corrective Action-Cost']
    for i in sheet:
        if i == '01Assembly':
            sh = work_book[i]
            col = 1
            while True:
                if col == 6:
                    break
                elif sh.cell(2, col).value == '[Part_No]':
                    sh.cell(2, col).value = part_number
                elif sh.cell(2, col).value == '[Part_Name]':
                    sh.cell(2, col).value = part_name
                elif sh.cell(2, col).value == '[Sys_Name]':
                    sh.cell(2, col).value = system_name
                col += 1
        elif i == '04Corrective Action-Cost':
            sh = work_book[i]
            row = 2
            while True:
                if re.match(r'.*\[Part_Name\].*', sh.cell(row, 1).value):
                    s = sh.cell(row, 1).value.replace('[Part_Name]', part_name)
                    sh.cell(row, 1).value = s
                if row == sh.max_row:
                    break
                row += 1
        else:
            sh = work_book[i]
            col = 4
            while True:
                if sh.cell(1, col).value == '输出':
                    break
                elif re.match(r'.*\[Part_Name\].*', sh.cell(1, col).value):
                    s = sh.cell(1, col).value.replace('[Part_Name]', part_name)
                    sh.cell(1, col).value = s
                col += 1
    save_file_name = str(part_number)+'_'+part_name+'_'+time.strftime("%Y%m%d%H%M%S", time.localtime())[:8]+'_Draft.xlsx'
    work_book.save(deal_data_path + '\\' + save_file_name)
    return send_from_directory(deal_data_path, save_file_name, as_attachment=True)


