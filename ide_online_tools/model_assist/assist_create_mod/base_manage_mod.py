# -*- coding:utf-8 -*-
import os, re
import time
from settings import Config
from openpyxl import load_workbook
from flask import request, send_from_directory, abort
from . import ac_mod


# 文件保存路径
initial_data_path = Config.ROOT_PATH + './data/ACM/initial_data'


def file_time(file):
    time_array = time.localtime(os.path.getmtime(file))
    other_time = time.strftime("%Y.%m.%d %H:%M:%S", time_array)[:10]
    return other_time


# ——————————————————————————————————————————————发送数据———————————————————————————
@ac_mod.route("/base_manage_mod/search_file", methods=['GET'])
def search_file():
    file_name_list = os.listdir(initial_data_path)
    table_data = []
    options_model_type = []
    options_model_number = []
    num = 1
    for x in file_name_list:
        if not re.match(r'~\$.*', x):
            model_type = x.rsplit('.', 1)[0]
            model_number = 'Model_000'+str(num)
            file_path = initial_data_path + "\\" + x
            modify_time = file_time(file_path)
            options_model_type.append({'value': model_type})
            options_model_number.append({'value': model_number})
            table_data.append({'modelNumber': model_number, 'modelType': model_type, 'modifyTime': modify_time, 'author': 'admin'})
            num += 1
    total_data = {'modelType': options_model_type, 'modelNumber': options_model_number, 'tableData': table_data}
    return total_data


# ——————————————————————————————————————————————删除文件———————————————————————————
@ac_mod.route("/base_manage_mod/delete_file", methods=['GET'])
def delete_file():
    file_name = request.args.get('modelType')
    file_path = initial_data_path + "\\" + file_name + '.xlsx'
    if os.path.exists(file_path):
        os.remove(file_path)
        return {"code": 1, "state": "success", "msg": '删除成功！'}
    else:
        abort(404)


# ——————————————————————————————————————————————上传数据———————————————————————————
@ac_mod.route("/base_manage_mod/upload_file", methods=['POST'])
def upload_file():
    file = request.files['file']
    work_book = load_workbook(file)
    model_type = work_book['01Assembly'].cell(2, 4).value
    if not model_type:
        model_type = '其他'
    file_name = str(model_type) + '.xlsx'
    work_book.save(os.path.join(initial_data_path, file_name))
    return {"code": 1, "state": "success", "msg": "上传成功！"}


# —————————————————————————————————————————————基础模型下载——————————————————————————
@ac_mod.route("/base_manage_mod/download_model", methods=['GET'])
def download_model():
    file_name = request.args.get('modelType') + '.xlsx'
    directory = initial_data_path
    return send_from_directory(directory, file_name, as_attachment=True)


# —————————————————————————————————————————————模板文件下载——————————————————————————
@ac_mod.route("/base_manage_mod/download_template", methods=['GET'])
def download_template():
    file_name = '其他.xlsx'
    directory = initial_data_path
    return send_from_directory(directory, file_name, as_attachment=True)


