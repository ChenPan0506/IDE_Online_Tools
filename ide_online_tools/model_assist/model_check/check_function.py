# coding:utf-8
# import os
# import sys
# from loguru import logger
# import re
# import numpy
import pandas as pd
from openpyxl.styles import PatternFill, colors
from openpyxl.utils import get_column_letter, column_index_from_string
import re

CHECK_SHEET_LIST = [
    "Model_Y",
    "01Assembly", "03FM Occurence Count", "04Corrective Action-Cost", "05FM-CA", "06Test", "07TestResult-Test",
    "Model", "Model_N"
]
P_DTC_NODE1 = re.compile('^[A-Z][A-Z]*[A-Z0-9]$')
P_DTC_NODE2 = re.compile('^[A-Z][A-Z0-9]*[A-Z0-9]$')
P_DTC_NODE3 = re.compile('^[A-Z][A-Z-]*[A-Z0-9]$')
P_DTC_NODE4 = re.compile('^[A-Z][A-Z0-9-]*[A-Z0-9]$')
FILL_WARNING = PatternFill(
    fill_type="solid",
    fgColor=colors.YELLOW
)
FILL_ERROR = PatternFill(
    fill_type="solid",
    fgColor=colors.RED
)


def when_info(msg_storage, msg):
    # logger.info(msg)
    msg_storage += (msg + "\r\n")
    return msg_storage


def when_warning(msg_storage, dict_warning, msg):
    # logger.warning(msg)
    msg_storage += (msg + "\r\n")
    dict_warning[len(dict_warning)] = msg
    return msg_storage, dict_warning


def when_error(msg_storage, dict_error, msg):
    # logger.error(msg)
    msg_storage += (msg + "\r\n")
    dict_error[len(dict_error)] = msg
    return msg_storage, dict_error


def check_part_num(func_stat, func_01, func_data, func_sheet, func_col, msg_storage, dict_warning, dict_error):
    if func_data.iloc[0, func_col] == "部件编号":
        for func_indx in range(1, func_data.shape[0]):
            if not pd.isnull(func_data.iloc[func_indx, func_col]):
                if func_data.iloc[func_indx, func_col] == func_01:
                    pass
                else:
                    func_stat = 0
                    msg_storage, dict_warning = when_warning(
                        msg_storage,
                        dict_warning,
                        CHECK_SHEET_LIST[func_sheet] + "中第" + str(func_indx + 1) + "行第" +
                        get_column_letter(func_col + 1) + "列存在异常，异常原因：'部件编号'项内容与" +
                        CHECK_SHEET_LIST[1] + "中不相匹")
            else:
                func_stat = 0
                msg_storage, dict_warning = when_warning(
                    msg_storage,
                    dict_warning,
                    CHECK_SHEET_LIST[func_sheet] + "中第" + str(func_indx + 1) + "行第" +
                    get_column_letter(func_col + 1) + "列存在异常，异常原因：'部件编号'项内容为空")
    else:
        func_stat = 0
        msg_storage, dict_error = when_error(
            msg_storage,
            dict_error,
            CHECK_SHEET_LIST[func_sheet] + "中第1行存在异常，异常原因：未找到名为'部件编号'的项")
    return func_stat, msg_storage
    # 校验部件编号


def check_part_name(func_stat, func_01, func_data, func_sheet, func_col, msg_storage, dict_warning, dict_error):
    if func_data.iloc[0, func_col] == "部件名称":
        for func_indx in range(1, func_data.shape[0]):
            if not pd.isnull(func_data.iloc[func_indx, func_col]):
                if func_data.iloc[func_indx, func_col] == func_01:
                    pass
                else:
                    func_stat = 0
                    msg_storage, dict_warning = when_warning(
                        msg_storage,
                        dict_warning,
                        CHECK_SHEET_LIST[func_sheet] + "中第" + str(func_indx + 1) + "行第" +
                        get_column_letter(func_col + 1) + "列存在异常，异常原因：'部件名称'项内容与" +
                        CHECK_SHEET_LIST[1] + "中不相匹")
            else:
                func_stat = 0
                msg_storage, dict_warning = when_warning(
                    msg_storage,
                    dict_warning,
                    CHECK_SHEET_LIST[func_sheet] + "中第" + str(func_indx + 1) + "行第" +
                    get_column_letter(func_col + 1) + "列存在异常，异常原因：'部件名称'项内容为空")
    else:
        func_stat = 0
        msg_storage, dict_error = when_error(
            msg_storage,
            dict_error,
            CHECK_SHEET_LIST[func_sheet] + "中第1行存在异常，异常原因：未找到名为'部件名称'的项")
    return func_stat, msg_storage
    # 校验部件名称


def check_sheet(odd_xl_data, str_xl_name, wb_xl):
    msg_storage = ""
    dict_warning = {}
    dict_error = {}

    list_sheet_name = list(odd_xl_data.keys())
    msg_storage = when_info(msg_storage, "====================开始校验EXCEL文件：" + str_xl_name + "====================")

    int_01_part_num_stat = 0
    int_01_part_name_stat = 0
    if CHECK_SHEET_LIST[1] in list_sheet_name:
        ws_01 = wb_xl[CHECK_SHEET_LIST[1]]

        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[1] + "表----------")
        df_01_data = odd_xl_data[CHECK_SHEET_LIST[1]]

        if df_01_data.iloc[0, 0] == "部件编号":
            if pd.isnull(df_01_data.iloc[1, 0]):
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[1] + "中第2行第A列存在异常，异常原因：'部件编号'项内容为空")
                ws_01.cell(row=2, column=1).fill = FILL_ERROR
            else:
                int_01_part_num_stat = 1
                int_01_part_num = df_01_data.iloc[1, 0]
                # logger.debug("'部件编号'项内容位于" + CHECK_SHEET_LIST[1] + "中第2行第1列，校验无误")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中第1行第A列存在异常，异常原因：未找到名为'部件编号'的项")
            ws_01.cell(row=1, column=1).fill = FILL_ERROR
            # 校验部件编号

        if df_01_data.iloc[0, 1] == "部件名称":
            if pd.isnull(df_01_data.iloc[1, 1]):
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[1] + "中第2行第B列存在异常，异常原因：'部件名称'项内容为空")
                ws_01.cell(row=2, column=2).fill = FILL_ERROR
            else:
                int_01_part_name_stat = 1
                str_01_part_name = df_01_data.iloc[1, 1]
                # logger.debug("'部件名称'项内容位于" + CHECK_SHEET_LIST[1] + "中第2行第2列，校验无误")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中第1行第B列存在异常，异常原因：未找到名为'部件名称'的项")
            ws_01.cell(row=1, column=2).fill = FILL_ERROR
    else:
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[1])
        # 校验部件名称
    stat_01 = int_01_part_num_stat and int_01_part_name_stat
    # 校验01Assembly表##################################################################################################

    int_fs_stat = 0
    int_fo_stat = 0
    int_is_stat = 0
    int_io_stat = 0
    int_mod_part_num_stat = 0
    int_mod_part_name_stat = 0
    int_mod_bf_part_name_stat = 1
    int_mod_fm_stat = 1
    int_mod_freq_stat = 1
    int_mod_test_stat = 1
    int_mod_input = 1
    int_mod_output = 1
    int_mod_area_fs_stat = 1
    int_mod_area_fo_stat = 1
    int_mod_area_is_stat = 1
    int_mod_area_io_stat = 1
    if CHECK_SHEET_LIST[0] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[0] + "表----------")
        df_mod_data = odd_xl_data[CHECK_SHEET_LIST[0]]
        ws_mod = wb_xl[CHECK_SHEET_LIST[0]]

        if df_mod_data.iloc[2, 2] == "FS":
            list_fs_loc = [2, 2]
            int_fs_stat = 1
            # logger.debug(
            #     "'FS'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 1) + "行第" + str(list_fs_loc[1] + 1) +
            #     "列，校验无误")
            indx_fo_loc = df_mod_data.iloc[list_fs_loc[0]][df_mod_data.iloc[list_fs_loc[0]].isin(["FO"])].index
            if indx_fo_loc.size == 1:
                list_fo_loc = [list_fs_loc[0], indx_fo_loc.tolist()[0]]
                int_fo_stat = 1
                # logger.debug(
                #     "'FO'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_fo_loc[0] + 1) + "行第" +
                #     str(list_fo_loc[1] + 1) + "列，校验无误")
            elif indx_fo_loc.size == 0:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 1) + "行存在异常，异常原因：未找到'FO'")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 1) + "行存在异常，异常原因：存在多个'FO'")
            indx_is_loc = df_mod_data[list_fs_loc[1]][df_mod_data[list_fs_loc[1]].isin(["IS"])].index
            if indx_is_loc.size == 1:
                list_is_loc = [indx_is_loc.tolist()[0], list_fs_loc[1]]
                int_is_stat = 1
                # logger.debug(
                #     "'IS'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_is_loc[0] + 1) + "行第" +
                #     str(list_is_loc[1] + 1) + "列，校验无误")
            elif indx_is_loc.size == 0:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[0] + "中第" + get_column_letter(list_fs_loc[1] + 1) +
                    "列存在异常，异常原因：未找到'FO'")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[0] + "中第" + get_column_letter(list_fs_loc[1] + 1) +
                    "列存在异常，异常原因：存在多个'FO'")
            if int_fo_stat and int_is_stat:
                if df_mod_data.iloc[list_is_loc[0], list_fo_loc[1]] == "IO":
                    list_io_loc = [list_is_loc[0], list_fo_loc[1]]
                    int_io_stat = 1
                    # logger.debug(
                    #     "'IO'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_io_loc[0] + 1) + "行第" +
                    #     str(list_io_loc[1] + 1) + "列，校验无误")
                else:
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[0] + "中第" + str(list_is_loc[0] + 1) + "行第" +
                        get_column_letter(list_fo_loc[1] + 1) + "列存在异常，异常原因：未找到'IO'")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    "'FO'、'IS'位置信息异常，无法获得'IO'位置信息")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[0] + "中第3行第C列存在异常，异常原因：未找到'FS'")
            ws_mod.cell(row=3, column=3).fill = FILL_ERROR
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'FS'位置信息异常，无法获得'FO'、'IS'、'IO'位置信息")
        # 检索并校验FS、FO、IS、IO索引位置

        if pd.isnull(df_mod_data.iloc[1, 0]):
            msg_storage, dict_warning = when_warning(
                msg_storage,
                dict_warning,
                CHECK_SHEET_LIST[0] + "中第2行第A列存在异常，异常原因：'部件编号'项内容为空")
            ws_mod.cell(row=2, column=1).fill = FILL_WARNING
            if int_01_part_num_stat:
                ws_mod.cell(row=2, column=1).value = int_01_part_num
        else:
            if int_01_part_num_stat:
                if df_mod_data.iloc[1, 0] == int_01_part_num:
                    int_mod_part_num_stat = 1
                    # logger.debug(
                    #     "'部件编号'项内容位于" + CHECK_SHEET_LIST[0] + "中第2行第1列，校验无误")
                else:
                    msg_storage, dict_warning = when_warning(
                        msg_storage,
                        dict_warning,
                        CHECK_SHEET_LIST[0] + "中第2行第A列存在异常，异常原因：'部件编号'项内容与" +
                        CHECK_SHEET_LIST[1] + "中相应内容不匹配")
                    ws_mod.cell(row=2, column=1).fill = FILL_WARNING
                    ws_mod.cell(row=2, column=1).value = int_01_part_num

            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[1] + "中'部件编号'信息未找到，无法完成校验")
        if pd.isnull(df_mod_data.iloc[1, 1]):
            msg_storage, dict_warning = when_warning(
                msg_storage,
                dict_warning,
                CHECK_SHEET_LIST[0] + "中第2行第B列存在异常，异常原因：'部件名称'项内容为空")
            ws_mod.cell(row=2, column=2).fill = FILL_WARNING
            if int_01_part_name_stat:
                ws_mod.cell(row=2, column=2).value = str_01_part_name
        else:
            if int_01_part_name_stat:
                if df_mod_data.iloc[1, 1] == str_01_part_name:
                    int_mod_part_name_stat = 1
                    # logger.debug(
                    #     "'部件名称'项内容位于" + CHECK_SHEET_LIST[0] + "中第2行第2列，校验无误")
                else:
                    msg_storage, dict_warning = when_warning(
                        msg_storage,
                        dict_warning,
                        CHECK_SHEET_LIST[0] + "中第2行第B列存在异常，异常原因：'部件名称'项内容与" +
                        CHECK_SHEET_LIST[1] + "中相应内容不匹配")
                    ws_mod.cell(row=2, column=2).fill = FILL_WARNING
                    ws_mod.cell(row=2, column=2).value = str_01_part_name
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[1] + "中'部件名称'信息未找到，无法完成校验")
        # 校验部件编号及部件名称

        if int_fs_stat and int_is_stat:
            for int_mod_bf_part_name_indx in range(list_fs_loc[0] + 1, list_is_loc[0]):
                if pd.isnull(df_mod_data.iloc[int_mod_bf_part_name_indx, 0]):
                    int_mod_bf_part_name_stat = 0
                    msg_storage, dict_warning = when_warning(
                        msg_storage,
                        dict_warning,
                        CHECK_SHEET_LIST[0] + "中第" + str(int_mod_bf_part_name_indx + 1) +
                        "行第A列存在异常，异常原因：'失效模式前的部件名称'内容为空")
                    ws_mod.cell(row=int_mod_bf_part_name_indx + 1, column=1).fill = FILL_WARNING
                    if int_01_part_name_stat:
                        ws_mod.cell(row=int_mod_bf_part_name_indx + 1, column=1).value = str_01_part_name
                else:
                    if int_01_part_name_stat:
                        if df_mod_data.iloc[int_mod_bf_part_name_indx, 0] == str_01_part_name:
                            continue
                        else:
                            int_mod_bf_part_name_stat = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[0] + "中第" + str(int_mod_bf_part_name_indx + 1) +
                                "行第A列存在异常，异常原因：'失效模式前的部件名称'项内容与" +
                                CHECK_SHEET_LIST[1] + "中'部件名称'不相匹")
                            ws_mod.cell(row=int_mod_bf_part_name_indx + 1, column=1).fill = FILL_WARNING
                            ws_mod.cell(row=int_mod_bf_part_name_indx + 1, column=1).value = str_01_part_name
                    else:
                        int_mod_bf_part_name_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[1] + "中'部件名称'信息未找到，无法校验'失效模式前的部件名称'")
        else:
            int_mod_bf_part_name_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'FS'、'IS'位置信息异常，无法获得'失效模式前的部件名称'位置信息")
        # if int_mod_bf_part_name_stat:
            # logger.debug(
            #     "'失效模式前的部件名称'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 2) + "到" +
            #     str(list_is_loc[0]) + "行第1列，校验无误")
        # 校验失效模式前的部件名称

        if int_fs_stat and int_is_stat:
            df_mod_fm_data = df_mod_data.iloc[list_fs_loc[0] + 1:list_is_loc[0], list_fs_loc[1] - 1]
            bool_mod_fm_duple = max(df_mod_fm_data.duplicated().values)
            if not bool_mod_fm_duple:
                for int_mod_fm_indx in range(list_fs_loc[0] + 1, list_is_loc[0]):
                    if pd.isnull(df_mod_data.iloc[int_mod_fm_indx, list_fs_loc[1] - 1]):
                        int_mod_fm_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[0] + "中第" + str(int_mod_fm_indx + 1) + "行第" +
                            get_column_letter(list_fs_loc[1]) + "列存在异常，异常原因：'失效模式'内容为空")
                        ws_mod.cell(row=int_mod_fm_indx + 1, column=list_fs_loc[1]).fill = FILL_ERROR
                # if int_mod_fm_stat:
                    # logger.debug(
                    #     "'失效模式'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 2) + "到" +
                    #     str(list_is_loc[0]) + "行第" + str(list_fs_loc[1]) + "列，校验无误")
            else:
                int_mod_fm_stat = 0
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 2) + "到" + str(list_is_loc[0]) + "行第" +
                    get_column_letter(list_fs_loc[1]) + "列存在异常，异常原因：'失效模式'存在重复项")
        else:
            int_mod_fm_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'FS'、'IS'位置信息异常，无法获得'失效模式'位置信息")
        # 校验失效模式

        if int_fs_stat and int_is_stat:
            for int_mod_freq_indx in range(list_fs_loc[0] + 1, list_is_loc[0]):
                if pd.isnull(df_mod_data.iloc[int_mod_freq_indx, list_fs_loc[1]]):
                    int_mod_freq_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[0] + "中第" + str(int_mod_freq_indx + 1) + "行第" +
                        get_column_letter(list_fs_loc[1] + 1) + "列存在异常，异常原因：'失效概率频度'内容为空")
                    ws_mod.cell(row=int_mod_freq_indx + 1, column=list_fs_loc[1] + 1).fill = FILL_ERROR
                else:
                    if isinstance(df_mod_data.iloc[int_mod_freq_indx, list_fs_loc[1]], int) or \
                            isinstance(df_mod_data.iloc[int_mod_freq_indx, list_fs_loc[1]], float):
                        if 0.1 <= df_mod_data.iloc[int_mod_freq_indx, list_fs_loc[1]] <= 10:
                            continue
                        else:
                            int_mod_freq_stat = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[0] + "中第" + str(int_mod_freq_indx + 1) + "行第" +
                                get_column_letter(list_fs_loc[1] + 1) +
                                "列存在异常，异常原因：内容无效，有效值为>=0.1且<=10的数字")
                            ws_mod.cell(row=int_mod_freq_indx + 1, column=list_fs_loc[1] + 1).fill = FILL_ERROR
                    else:
                        int_mod_freq_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[0] + "中第" + str(int_mod_freq_indx + 1) + "行第" +
                            get_column_letter(list_fs_loc[1] + 1) +
                            "列存在异常，异常原因：内容无效，有效值为>=0.1且<=10的数字")
                        ws_mod.cell(row=int_mod_freq_indx + 1, column=list_fs_loc[1] + 1).fill = FILL_ERROR
            # if int_mod_freq_stat:
                # logger.debug(
                #     "'失效概率频度'位于" + CHECK_SHEET_LIST[0] + "中第" + str(list_fs_loc[0] + 2) + "到" +
                #     str(list_is_loc[0]) + "行第" + str(list_fs_loc[1] + 1) + "列，校验无误")
        else:
            int_mod_freq_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'FS'、'IS'位置信息异常，无法获得'失效概率频度'位置信息")
        # 校验失效概率频度

        if int_fs_stat and int_fo_stat:
            list_mod_test = []

            for int_mod_test_col in range(list_fs_loc[1] + 1, list_fo_loc[1]):
                if str(df_mod_data.iloc[0, int_mod_test_col]).startswith("系统症状"):
                    continue
                elif str(df_mod_data.iloc[0, int_mod_test_col]).startswith("DTC_"):
                    targ_str = str(df_mod_data.iloc[0, int_mod_test_col])
                    targ_list = re.split("_", targ_str)
                    targ_str_stat = 0

                    if len(targ_list) == 3:
                        if targ_list[0] == "DTC" and len(targ_list[2]) == 7:
                            if P_DTC_NODE1.match(targ_list[1]) is not None or \
                                    P_DTC_NODE2.match(targ_list[1]) is not None or \
                                    P_DTC_NODE3.match(targ_list[1]) is not None or \
                                    P_DTC_NODE4.match(targ_list[1]) is not None:
                                targ_str_stat = 1
                    if targ_str_stat:
                        pass
                    else:
                        int_mod_test_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[0] + "中第1行第" + get_column_letter(int_mod_test_col + 1) +
                            "列存在异常，异常原因：'DTC'命名不规范")
                        ws_mod.cell(row=1, column=int_mod_test_col + 1).fill = FILL_ERROR
                elif str(df_mod_data.iloc[0, int_mod_test_col]).startswith("部件症状_"):
                    if int_01_part_name_stat:
                        if df_mod_data.iloc[0, int_mod_test_col] == "部件症状_" + str_01_part_name:
                            if pd.isnull(df_mod_data.iloc[1, int_mod_test_col]):
                                int_mod_test_stat = 0
                                msg_storage, dict_error = when_error(
                                    msg_storage,
                                    dict_error,
                                    CHECK_SHEET_LIST[0] + "中第2行第" + get_column_letter(int_mod_test_col + 1) +
                                    "列存在异常，异常原因：'部件症状'项内容为空")
                                ws_mod.cell(row=2, column=int_mod_test_col + 1).fill = FILL_ERROR
                        else:
                            int_mod_test_stat = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[0] + "中第1行第" + get_column_letter(int_mod_test_col + 1) +
                                "列存在异常，异常原因：'部件症状'命名不规范")
                            ws_mod.cell(row=1, column=int_mod_test_col + 1).fill = FILL_WARNING
                            ws_mod.cell(row=1, column=int_mod_test_col + 1).value = "部件症状_" + str_01_part_name
                    else:
                        int_mod_test_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[1] + "中'部件名称'信息未找到，无法校验'部件症状'")
                elif str(df_mod_data.iloc[0, int_mod_test_col]).startswith("Test_"):
                    if int_01_part_name_stat:
                        if df_mod_data.iloc[0, int_mod_test_col] == "Test_" + str_01_part_name:
                            if pd.isnull(df_mod_data.iloc[1, int_mod_test_col]):
                                int_mod_test_stat = 0
                                msg_storage, dict_error = when_error(
                                    msg_storage,
                                    dict_error,
                                    CHECK_SHEET_LIST[0] + "中第2行第" + get_column_letter(int_mod_test_col + 1) +
                                    "列存在异常，异常原因：'测试异常'项内容为空")
                                ws_mod.cell(row=2, column=int_mod_test_col + 1).fill = FILL_ERROR
                            else:
                                list_mod_test.append(df_mod_data.iloc[1, int_mod_test_col])
                        else:
                            int_mod_test_stat = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[0] + "中第1行第" + get_column_letter(int_mod_test_col + 1) +
                                "列存在异常，异常原因：'测试异常'命名不规范")
                            ws_mod.cell(row=1, column=int_mod_test_col + 1).fill = FILL_WARNING
                            ws_mod.cell(row=1, column=int_mod_test_col + 1).value = "Test_" + str_01_part_name
                    else:
                        int_mod_test_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[1] + "中'部件名称'信息未找到，无法校验'测试异常'")
                else:
                    int_mod_test_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[0] + "中第1行第" + get_column_letter(int_mod_test_col + 1) +
                        "列存在异常，异常原因：无法识别该症状类别")
                    ws_mod.cell(row=1, column=int_mod_test_col + 1).fill = FILL_ERROR
        else:
            int_mod_test_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'FS'、'FO'位置信息异常，无法获得'症状'位置信息")
        # if int_mod_test_stat:
            # logger.debug("'症状'位于" + CHECK_SHEET_LIST[0] + "中第1、2行，校验无误")
        # 校验症状

        if int_is_stat:
            for int_mod_input_indx in range(list_is_loc[0] + 1, df_mod_data.shape[0]):
                for int_mod_input_col in range(2):
                    if pd.isnull(df_mod_data.iloc[int_mod_input_indx, int_mod_input_col]):
                        int_mod_input = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[0] + "中第" + str(int_mod_input_indx + 1) + "行第" +
                            get_column_letter(int_mod_input_col) + "列存在异常，异常原因：'输入及接口异常'内容为空")
                        ws_mod.cell(row=int_mod_input_indx + 1, column=int_mod_input_col + 1).fill = FILL_ERROR
        else:
            int_mod_input = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'IS'位置信息异常，无法获得'输入及接口异常'位置信息")
        # if int_mod_input:
            # logger.debug(
            #     "'输入及接口异常'位于" + CHECK_SHEET_LIST[0] + "中第" +
            #     str(list_is_loc[0] + 2) + "到" + str(df_mod_data.shape[0]) + "行第1到2列，校验无误")
        # 校验输入及接口异常

        if int_fo_stat:
            for int_mod_output_indx in range(2):
                for int_mod_output_col in range(list_fo_loc[1] + 1, df_mod_data.shape[1]):
                    if pd.isnull(df_mod_data.iloc[int_mod_output_indx, int_mod_output_col]):
                        int_mod_output = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[0] + "中第" + str(int_mod_output_indx + 1) + "行第" +
                            get_column_letter(int_mod_output_col) + "列存在异常，异常原因：'输出及接口异常'内容为空")
                        ws_mod.cell(row=int_mod_output_indx + 1, column=int_mod_output_col + 1).fill = FILL_ERROR
        else:
            int_mod_output = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'FO'位置信息异常，无法获得'输出及接口异常'位置信息")
        # if int_mod_output:
            # logger.debug(
            #     "'输出及接口异常'位于" + CHECK_SHEET_LIST[0] + "中第1到2行第" + str(list_fo_loc[1] + 2) +
            #     "到" + str(df_mod_data.shape[1]) + "列，校验无误")
        # 校验输出及接口异常

        if int_io_stat:
            for int_mod_area_fs_indx in range(list_fs_loc[0] + 1, list_io_loc[0]):
                int_mod_area_fs_row_stat = 0
                # for int_mod_area_fs_col in range(list_fs_loc[1] + 1, df_mod_n_data.shape[0]):
                for int_mod_area_fs_col in range(list_fs_loc[1] + 1, df_mod_data.shape[1]):
                    if pd.isnull(df_mod_data.iloc[int_mod_area_fs_indx, int_mod_area_fs_col]):
                        pass
                    else:
                        int_mod_area_fs_row_stat = 1
                        if df_mod_data.iloc[int_mod_area_fs_indx, int_mod_area_fs_col] == "X":
                            pass
                        else:
                            int_mod_area_fs_stat = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[0] + "中第" + str(int_mod_area_fs_indx + 1) + "行第" +
                                get_column_letter(int_mod_area_fs_col + 1) + "列存在异常，异常原因：内容无效，有效值为'X'")
                            ws_mod.cell(
                                row=int_mod_area_fs_indx + 1, column=int_mod_area_fs_col + 1).fill = FILL_WARNING
                            ws_mod.cell(
                                row=int_mod_area_fs_indx + 1, column=int_mod_area_fs_col + 1).value = "X"
                if not int_mod_area_fs_row_stat:
                    int_mod_area_fs_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[0] + "中第" + str(int_mod_area_fs_indx + 1) + "行第" +
                        get_column_letter(list_fs_loc[1] + 2) + "到" + get_column_letter(list_io_loc[1]) +
                        "列存在异常，异常原因：未找到该行'失效模式-症状'信息")
        else:
            int_mod_area_fs_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'IO'位置信息异常，无法获得'失效模式-症状'位置信息")
        # if int_mod_area_fs_stat:
            # logger.debug(
            #     "'失效模式-症状'位于" + CHECK_SHEET_LIST[0] + "中第" +
            #     str(list_fs_loc[0] + 2) + "到" + str(list_io_loc[0]) + "行第" +
            #     str(list_fs_loc[1] + 2) + "到" + str(list_io_loc[1]) + "列，校验无误")
        # 校验失效模式-症状

        if int_io_stat:
            for int_mod_area_fo_indx in range(list_fs_loc[0] + 1, list_io_loc[0]):
                for int_mod_area_fo_col in range(list_io_loc[1] + 1, df_mod_data.shape[1]):
                    if pd.isnull(df_mod_data.iloc[int_mod_area_fo_indx, int_mod_area_fo_col]):
                        continue
                    else:
                        if df_mod_data.iloc[int_mod_area_fo_indx, int_mod_area_fo_col] == "X":
                            continue
                        else:
                            int_mod_area_fo_stat = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[0] + "中第" + str(int_mod_area_fo_indx + 1) + "行第" +
                                get_column_letter(int_mod_area_fo_col + 1) + "列存在异常，异常原因：内容无效，有效值为'X'")
                            ws_mod.cell(
                                row=int_mod_area_fo_indx + 1, column=int_mod_area_fo_col + 1).fill = FILL_WARNING
                            ws_mod.cell(
                                row=int_mod_area_fo_indx + 1, column=int_mod_area_fo_col + 1).value = "X"
        else:
            int_mod_area_fo_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'IO'位置信息异常，无法获得'失效模式-输出'位置信息")
        # if int_mod_area_fo_stat:
            # logger.debug(
            #     "'失效模式-输出'位于" + CHECK_SHEET_LIST[0] + "中第" +
            #     str(list_fs_loc[0] + 2) + "到" + str(list_io_loc[0]) + "行第" +
            #     str(list_io_loc[1] + 2) + "到" + str(df_mod_data.shape[1]) + "列，校验无误")
        # 校验失效模式-输出

        if int_io_stat:
            for int_mod_area_is_indx in range(list_io_loc[0] + 1, df_mod_data.shape[0]):
                for int_mod_area_is_col in range(list_fs_loc[1] + 1, list_io_loc[1]):
                    if pd.isnull(df_mod_data.iloc[int_mod_area_is_indx, int_mod_area_is_col]):
                        continue
                    else:
                        if df_mod_data.iloc[int_mod_area_is_indx, int_mod_area_is_col] == "X":
                            continue
                        else:
                            int_mod_area_is_stat = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[0] + "中第" + str(int_mod_area_is_indx + 1) + "行第" +
                                get_column_letter(int_mod_area_is_col + 1) + "列存在异常，异常原因：内容无效，有效值为'X'")
                            ws_mod.cell(
                                row=int_mod_area_is_indx + 1, column=int_mod_area_is_col + 1).fill = FILL_WARNING
                            ws_mod.cell(
                                row=int_mod_area_is_indx + 1, column=int_mod_area_is_col + 1).value = "X"
        else:
            int_mod_area_is_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'IO'位置信息异常，无法获得'输入-症状'位置信息")
        # if int_mod_area_is_stat:
            # logger.debug(
            #     "'输入-症状'位于" + CHECK_SHEET_LIST[0] + "中第" +
            #     str(list_io_loc[0] + 2) + "到" + str(df_mod_data.shape[0]) + "行第" +
            #     str(list_fs_loc[1] + 2) + "到" + str(list_io_loc[1]) + "列，校验无误")
        # 校验输入-症状

        if int_mod_input and int_mod_output:
            if int_io_stat:
                for int_mod_area_io_indx in range(list_io_loc[0] + 1, df_mod_data.shape[0]):
                    for int_mod_area_io_col in range(list_io_loc[1] + 1, df_mod_data.shape[1]):
                        if not pd.isnull(df_mod_data.iloc[int_mod_area_io_indx, int_mod_area_io_col]):
                            if df_mod_data.iloc[int_mod_area_io_indx, int_mod_area_io_col] == "X":
                                pass
                            else:
                                int_mod_area_io_stat = 0
                                msg_storage, dict_warning = when_warning(
                                    msg_storage,
                                    dict_warning,
                                    CHECK_SHEET_LIST[0] + "中第" + str(int_mod_area_io_indx + 1) + "行第" +
                                    get_column_letter(int_mod_area_io_col + 1) + "列存在异常，异常原因：内容无效，有效值为'X'")
                                ws_mod.cell(
                                    row=int_mod_area_io_indx + 1, column=int_mod_area_io_col + 1).fill = FILL_WARNING
                                ws_mod.cell(
                                    row=int_mod_area_io_indx + 1, column=int_mod_area_io_col + 1).value = "X"
                        if df_mod_data.iloc[int_mod_area_io_indx, 0] == df_mod_data.iloc[0, int_mod_area_io_col] and \
                                df_mod_data.iloc[int_mod_area_io_indx, 1] == df_mod_data.iloc[1, int_mod_area_io_col]:
                            if df_mod_data.iloc[int_mod_area_io_indx, int_mod_area_io_col] != "X":
                                int_mod_area_io_stat = 0
                                msg_storage, dict_warning = when_warning(
                                    msg_storage,
                                    dict_warning,
                                    CHECK_SHEET_LIST[0] + "中第" + str(int_mod_area_io_indx + 1) + "行第" +
                                    get_column_letter(int_mod_area_io_col + 1) + "列存在异常，异常原因：内容无效，有效值为X")
                                ws_mod.cell(
                                    row=int_mod_area_io_indx + 1, column=int_mod_area_io_col + 1).fill = FILL_WARNING
                                ws_mod.cell(
                                    row=int_mod_area_io_indx + 1, column=int_mod_area_io_col + 1).value = "X"
            else:
                int_mod_area_io_stat = 0
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    "'IO'位置信息异常，无法获得'输入-输出'位置信息")
        else:
            int_mod_area_io_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "'输入及接口异常'、'输出及接口异常'位置信息异常，无法校验'输入-输出'信息")
        # if int_mod_area_io_stat:
            # logger.debug(
            #     "'输入-输出'位于" + CHECK_SHEET_LIST[0] + "中第" +
            #     str(list_io_loc[0] + 2) + "到" + str(df_mod_data.shape[0]) + "行第" +
            #     str(list_io_loc[1] + 2) + "到" + str(df_mod_data.shape[1]) + "列，校验无误")
        # 校验输入-输出
    else:
        int_mod_bf_part_name_stat = 0
        int_mod_fm_stat = 0
        int_mod_freq_stat = 0
        int_mod_test_stat = 0
        int_mod_input = 0
        int_mod_output = 0
        int_mod_area_fs_stat = 0
        int_mod_area_fo_stat = 0
        int_mod_area_is_stat = 0
        int_mod_area_io_stat = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[0])
    stat_mod = \
        int_fs_stat and int_fo_stat and int_is_stat and int_io_stat and \
        int_mod_part_num_stat and int_mod_part_name_stat and int_mod_bf_part_name_stat and \
        int_mod_fm_stat and int_mod_freq_stat and int_mod_test_stat and int_mod_input and int_mod_output and \
        int_mod_area_fs_stat and int_mod_area_fo_stat and int_mod_area_is_stat and int_mod_area_io_stat
    # 校验Model_Y表#######################################################################################################

    int_03_fm_freq_stat = 0
    int_03_part_num_stat = 1
    int_03_part_name_stat = 1
    if CHECK_SHEET_LIST[2] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[2] + "表----------")
        df_03_data = odd_xl_data[CHECK_SHEET_LIST[2]]

        if int_mod_fm_stat and int_mod_freq_stat:
            if df_03_data.iloc[0, 0] == "失效模式" and df_03_data.iloc[0, 1] == "失效概率频度系数":
                dict_mod_fm_freq = df_mod_data.iloc[
                                   list_fs_loc[0] + 1:list_is_loc[0], list_fs_loc[1] - 1:list_fs_loc[1] + 1].set_index(
                    list_fs_loc[1] - 1).T.to_dict('list')
                dict_03_fm_freq = df_03_data.iloc[1:, 0:2].set_index(0).T.to_dict('list')
                if dict_mod_fm_freq == dict_03_fm_freq:
                    int_03_fm_freq_stat = 1
                    # logger.debug(
                    #     "'失效模式'和'失效概率频度'信息位于" + CHECK_SHEET_LIST[2] + "中第2到" +
                    #     str(df_03_data.shape[0]) + "行第1到2列，校验无误")
                else:
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[2] + "中第2到" + str(df_03_data.shape[0]) +
                        "行第A到B列存在异常，异常原因：'失效模式'和'失效概率频度'信息与" + CHECK_SHEET_LIST[0] +
                        "中不相匹")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[2] + "中第1行存在异常，异常原因：未找到名为'失效模式'或'失效概率频度系数'的项")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[0] + "中'失效模式'和'失效概率频度'信息异常，无法完成'失效模式-失效概率频度'校验")
        # 校验失效模式-失效概率频度

        if int_01_part_num_stat:
            int_03_part_num_stat, msg_storage = \
                check_part_num(
                    int_03_part_num_stat, int_01_part_num, df_03_data, 2, 3, msg_storage, dict_warning, dict_error)
        else:
            int_03_part_num_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中'部件编号'信息异常，无法完成'部件编号'校验")
        # if int_03_part_num_stat:
            # logger.debug(
            #     "'部件编号'信息位于" + CHECK_SHEET_LIST[2] + "中第2到" + str(df_03_data.shape[0]) + "行第4列，校验无误")
        # 校验部件编号

        if int_01_part_name_stat:
            int_03_part_name_stat, msg_storage = \
                check_part_name(
                    int_03_part_name_stat, str_01_part_name, df_03_data, 2, 4, msg_storage, dict_warning, dict_error)
        else:
            int_03_part_name_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中'部件名称'信息异常，无法完成'部件名称'校验")
        # if int_03_part_name_stat:
            # logger.debug(
            #     "'部件名称'信息位于" + CHECK_SHEET_LIST[2] + "中第2到" + str(df_03_data.shape[0]) + "行第5列，校验无误")
        # 校验部件名称
    else:
        int_03_part_num_stat = 0
        int_03_part_name_stat = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[2])
    stat_03 = int_03_fm_freq_stat and int_03_part_num_stat and int_03_part_name_stat
    # 校验03FM Occurence Count表########################################################################################

    int_04_fix_stat = 1
    int_04_fix_time_stat = 1
    int_04_fix_cost_stat = 1
    int_04_part_num_stat = 1
    int_04_part_name_stat = 1
    if CHECK_SHEET_LIST[3] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[3] + "表----------")
        df_04_data = odd_xl_data[CHECK_SHEET_LIST[3]]

        if df_04_data.iloc[0, 0] == "维修措施":
            for int_04_fix_indx in range(1, df_04_data.shape[0]):
                if pd.isnull(df_04_data.iloc[int_04_fix_indx, 0]):
                    int_04_fix_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[3] + "中第" + str(int_04_fix_indx + 1) +
                        "行第A列存在异常，异常原因：'维修措施'项内容为空")
        else:
            int_04_fix_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[3] + "中第1行存在异常，异常原因：未找到名为'维修措施'的项")
        # if int_04_fix_stat:
            # logger.debug(
            #     "'维修措施'信息位于" + CHECK_SHEET_LIST[3] + "中第2到" + str(df_04_data.shape[0]) + "行第1列，校验无误")
        # 校验维修措施

        if df_04_data.iloc[0, 1] == "工时（h）":
            for int_04_fix_time_indx in range(1, df_04_data.shape[0]):
                if not pd.isnull(df_04_data.iloc[int_04_fix_time_indx, 1]):
                    if isinstance(df_04_data.iloc[int_04_fix_time_indx, 1], int) or \
                            isinstance(df_04_data.iloc[int_04_fix_time_indx, 1], float):
                        pass
                    else:
                        int_04_fix_time_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[3] + "中第" + str(int_04_fix_time_indx + 1) +
                            "行第B列存在异常，异常原因：'工时（h）'项内容不合规")
                else:
                    int_04_fix_time_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[3] + "中第" + str(int_04_fix_time_indx + 1) +
                        "行第B列存在异常，异常原因：'工时（h）'项内容为空")
        else:
            int_04_fix_time_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[3] + "中第1行存在异常，异常原因：未找到名为'工时（h）'的项")
        # if int_04_fix_time_stat:
            # logger.debug(
            #     "'工时（h）'信息位于" + CHECK_SHEET_LIST[3] + "中第2到" + str(df_04_data.shape[0]) + "行第2列，校验无误")
        # 校验工时

        if df_04_data.iloc[0, 2] == "成本（元）":
            for int_04_fix_cost_indx in range(1, df_04_data.shape[0]):
                if not pd.isnull(df_04_data.iloc[int_04_fix_cost_indx, 2]):
                    if isinstance(df_04_data.iloc[int_04_fix_cost_indx, 2], int) or \
                            isinstance(df_04_data.iloc[int_04_fix_cost_indx, 2], float):
                        pass
                    else:
                        int_04_fix_cost_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[3] + "中第" + str(int_04_fix_cost_indx + 1) +
                            "行第C列存在异常，异常原因：'成本（元）'项内容不合规")
                else:
                    int_04_fix_cost_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[3] + "中第" + str(int_04_fix_cost_indx + 1) +
                        "行第C列存在异常，异常原因：'成本（元）'项内容为空")
        else:
            int_04_fix_cost_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[3] + "中第1行存在异常，异常原因：未找到名为'成本（元）'的项")
        # if int_04_fix_cost_stat:
            # logger.debug(
            #     "'成本（元）'信息位于" + CHECK_SHEET_LIST[3] + "中第2到" + str(df_04_data.shape[0]) + "行第3列，校验无误")
        # 校验成本

        if int_01_part_num_stat:
            int_04_part_num_stat, msg_storage = \
                check_part_num(
                    int_04_part_num_stat, int_01_part_num, df_04_data, 3, 3, msg_storage, dict_warning, dict_error)
        else:
            int_04_part_num_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中'部件编号'信息异常，无法完成'部件编号'校验")
        # if int_04_part_num_stat:
            # logger.debug(
            #     "'部件编号'信息位于" + CHECK_SHEET_LIST[3] + "中第2到" + str(df_04_data.shape[0]) + "行第4列，校验无误")
        # 校验部件编号

        if int_01_part_name_stat:
            int_04_part_name_stat, msg_storage = \
                check_part_name(
                    int_04_part_name_stat, str_01_part_name, df_04_data, 3, 4, msg_storage, dict_warning, dict_error)
        else:
            int_04_part_name_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中'部件名称'信息异常，无法完成'部件名称'校验")
        # if int_04_part_name_stat:
            # logger.debug(
            #     "'部件名称'信息位于" + CHECK_SHEET_LIST[3] + "中第2到" + str(df_04_data.shape[0]) + "行第5列，校验无误")
        # 校验部件名称
    else:
        int_04_fix_stat = 0
        int_04_fix_time_stat = 0
        int_04_fix_cost_stat = 0
        int_04_part_num_stat = 0
        int_04_part_name_stat = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[3])
    stat_04 = \
        int_04_fix_stat and int_04_fix_time_stat and int_04_fix_cost_stat and \
        int_04_part_num_stat and int_04_part_name_stat
    # 校验04Corrective Action-Cost######################################################################################

    int_05_fm_stat = 0
    int_05_fix_stat = 0
    int_05_area_fm_fix_stat = 1
    if CHECK_SHEET_LIST[4] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[4] + "表----------")
        df_05_data = odd_xl_data[CHECK_SHEET_LIST[4]]

        if int_mod_fm_stat:
            df_05_fm_data = df_05_data.iloc[1:, 0]
            bool_05_fm_duple = max(df_05_fm_data.duplicated().values)
            if not bool_05_fm_duple:
                if set(df_05_fm_data) == set(df_mod_fm_data):
                    int_05_fm_stat = 1
                    # logger.debug(
                    #     "'失效模式'位于" + CHECK_SHEET_LIST[4] + "中第1到" + str(df_05_data.shape[0]) +
                    #     "行第1列，校验无误")
                else:
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[4] + "中第2到" + str(df_05_data.shape[0]) +
                        "行第A列存在异常，异常原因：'失效模式'项内容与" + CHECK_SHEET_LIST[0] + "中不相匹")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[4] + "中第2到" + str(df_05_data.shape[0]) +
                    "行第A列存在异常，异常原因：'失效模式'存在重复项")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[0] + "中'失效模式'信息异常，无法完成'失效模式'校验")
        # 校验失效模式

        if int_04_fix_stat:
            df_05_fix_data = df_05_data.iloc[0, 1:]
            if set(df_05_fix_data) == set(df_04_data.iloc[1:, 0]):
                int_05_fix_stat = 1
                # logger.debug(
                #     "'维修措施'位于" + CHECK_SHEET_LIST[4] + "中第1行第2到" + str(df_05_data.shape[1]) + "列，校验无误")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[4] + "中第1行第2到" + get_column_letter(df_05_data.shape[1]) +
                    "列存在异常，异常原因：'维修措施'项内容与" + CHECK_SHEET_LIST[3] + "中不相匹")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[3] + "中'维修措施'信息异常，无法完成'维修措施'校验")
        # 校验维修措施

        if int_05_fm_stat and int_05_fix_stat:
            for int_05_area_fm_fix_indx in range(1, df_05_data.shape[0]):
                int_05_row_fm_fix_stat = 0
                for int_05_area_fm_fix_col in range(1, df_05_data.shape[1]):
                    if pd.isnull(df_05_data.iloc[int_05_area_fm_fix_indx, int_05_area_fm_fix_col]):
                        pass
                    else:
                        int_05_row_fm_fix_stat = 1
                        if df_05_data.iloc[int_05_area_fm_fix_indx, int_05_area_fm_fix_col] == "X" or \
                                df_05_data.iloc[int_05_area_fm_fix_indx, int_05_area_fm_fix_col] == 1:
                            pass
                        else:
                            int_05_area_fm_fix_stat = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[4] + "中第" + str(int_05_area_fm_fix_indx + 1) + "行第" +
                                get_column_letter(int_05_area_fm_fix_col + 1) +
                                "列存在异常，异常原因：内容无效，有效值为'X'或1")
                if not int_05_row_fm_fix_stat:
                    int_05_area_fm_fix_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[4] + "中第" + str(int_05_area_fm_fix_indx + 1) + "行第2到" +
                        get_column_letter(df_05_data.shape[1]) +
                        "列存在异常，异常原因：未找到该行'失效模式-维修措施'信息")
        else:
            int_05_area_fm_fix_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[4] + "中'失效模式'、'维修措施'信息异常，无法完成'失效模式-维修措施'校验")
        # if int_05_area_fm_fix_stat:
            # logger.debug(
            #     "'失效模式-维修措施'位于" + CHECK_SHEET_LIST[4] + "中第2到" + str(df_05_data.shape[0]) + "行第2到" +
            #     str(df_05_data.shape[1]) + "列，校验无误")
        # 校验失效模式-维修措施
    else:
        int_05_area_fm_fix_stat = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[4])
    stat_05 = int_05_fm_stat and int_05_fix_stat and int_05_area_fm_fix_stat
    # 校验05FM-CA#######################################################################################################

    int_06_test_type_stat = 1
    int_06_test_time_stat = 1
    int_06_test_cost_stat = 1
    int_06_part_num_stat = 1
    int_06_part_name_stat = 1
    if CHECK_SHEET_LIST[5] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[5] + "表----------")
        df_06_data = odd_xl_data[CHECK_SHEET_LIST[5]]

        if df_06_data.iloc[0, 0] == "测试项目":
            for int_06_test_type_indx in range(1, df_06_data.shape[0]):
                if pd.isnull(df_06_data.iloc[int_06_test_type_indx, 0]):
                    int_06_test_type_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[5] + "中第" + str(int_06_test_type_indx + 1) +
                        "行第A列存在异常，异常原因：'测试项目'项内容为空")
        else:
            int_06_test_type_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[5] + "中第1行存在异常，异常原因：未找到名为'测试项目'的项")
        # if int_06_test_type_stat:
            # logger.debug(
            #     "'测试项目'信息位于" + CHECK_SHEET_LIST[5] + "中第2到" + str(df_06_data.shape[0]) + "行第1列，校验无误")
        # 校验测试项目

        if df_06_data.iloc[0, 1] == "工时（h）":
            for int_06_test_time_indx in range(1, df_06_data.shape[0]):
                if not pd.isnull(df_06_data.iloc[int_06_test_time_indx, 1]):
                    if isinstance(df_06_data.iloc[int_06_test_time_indx, 1], int) or \
                            isinstance(df_06_data.iloc[int_06_test_time_indx, 1], float):
                        pass
                    else:
                        int_06_test_time_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[5] + "中第" + str(int_06_test_time_indx + 1) +
                            "行第B列存在异常，异常原因：'工时（h）'项内容不合规")
                else:
                    int_06_test_time_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[5] + "中第" + str(int_06_test_time_indx + 1) +
                        "行第B列存在异常，异常原因：'工时（h）'项内容为空")
        else:
            int_06_test_time_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[5] + "中第1行存在异常，异常原因：未找到名为'工时（h）'的项")
        # if int_06_test_time_stat:
            # logger.debug(
            #     "'工时（h）'信息位于" + CHECK_SHEET_LIST[5] + "中第2到" + str(df_06_data.shape[0]) + "行第2列，校验无误")
        # 校验工时

        if df_06_data.iloc[0, 2] == "成本（元）":
            for int_06_test_cost_indx in range(1, df_06_data.shape[0]):
                if not pd.isnull(df_06_data.iloc[int_06_test_cost_indx, 2]):
                    if isinstance(df_06_data.iloc[int_06_test_cost_indx, 2], int) or \
                            isinstance(df_06_data.iloc[int_06_test_cost_indx, 2], float):
                        pass
                    else:
                        int_06_test_cost_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[5] + "中第" + str(int_06_test_cost_indx + 1) +
                            "行第C列存在异常，异常原因：'成本（元）'项内容不合规")
                else:
                    int_06_test_cost_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[5] + "中第" + str(int_06_test_cost_indx + 1) +
                        "行第C列存在异常，异常原因：'成本（元）'项内容为空")
        else:
            int_06_test_cost_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[5] + "中第1行存在异常，异常原因：未找到名为'成本（元）'的项")
        # if int_06_test_cost_stat:
            # logger.debug(
            #     "'成本（元）'信息位于" + CHECK_SHEET_LIST[5] + "中第2到" + str(df_06_data.shape[0]) + "行第3列，校验无误")
        # 校验成本

        if int_01_part_num_stat:
            int_06_part_num_stat, msg_storage = \
                check_part_num(
                    int_06_part_num_stat, int_01_part_num, df_06_data, 5, 3, msg_storage, dict_warning, dict_error)
        else:
            int_06_part_num_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中'部件编号'信息异常，无法完成'部件编号'校验")
        # if int_06_part_num_stat:
            # logger.debug(
            #     "'部件编号'信息位于" + CHECK_SHEET_LIST[5] + "中第2到" + str(df_06_data.shape[0]) + "行第4列，校验无误")
        # 校验部件编号

        if int_01_part_name_stat:
            int_06_part_name_stat, msg_storage = \
                check_part_name(
                    int_06_part_name_stat, str_01_part_name, df_06_data, 5, 4, msg_storage, dict_warning, dict_error)
        else:
            int_06_part_name_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[1] + "中'部件名称'信息异常，无法完成'部件名称'校验")
        # if int_06_part_name_stat:
            # logger.debug(
            #     "'部件名称'信息位于" + CHECK_SHEET_LIST[5] + "中第2到" + str(df_06_data.shape[0]) + "行第5列，校验无误")
        # 校验部件名称
    else:
        int_06_test_type_stat = 0
        int_06_test_time_stat = 0
        int_06_test_cost_stat = 0
        int_06_part_num_stat = 0
        int_06_part_name_stat = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[5])
    stat_06 = \
        int_06_test_type_stat and int_06_test_time_stat and int_06_test_cost_stat and \
        int_06_part_num_stat and int_06_part_name_stat
    # 校验06Test########################################################################################################

    int_07_test_result_stat = 1
    int_07_result_num_stat = 1
    int_07_test_type_stat = 0
    int_07_area_result_type_stat = 1
    int_07_link_result_num_stat = 1
    int_07_link_num_type_stat = 1
    if CHECK_SHEET_LIST[6] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[6] + "表----------")
        df_07_data = odd_xl_data[CHECK_SHEET_LIST[6]]

        if df_07_data.iloc[0, 0] == "测试结果":
            for int_07_test_result_indx in range(1, df_07_data.shape[0]):
                if pd.isnull(df_07_data.iloc[int_07_test_result_indx, 0]):
                    int_07_test_result_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[6] + "中第" + str(int_07_test_result_indx + 1) +
                        "行第A列存在异常，异常原因：'测试项目'项内容为空")
        else:
            int_07_test_result_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[6] + "中第1行存在异常，异常原因：未找到名为'测试结果'的项")
        # if int_07_test_result_stat:
            # logger.debug(
            #     "'测试结果'信息位于" + CHECK_SHEET_LIST[6] + "中第2到" + str(df_07_data.shape[0]) + "行第1列，校验无误")
        # 校验测试结果

        if df_07_data.iloc[0, 1] == "结果判定":
            for int_07_result_num_indx in range(1, df_07_data.shape[0]):
                if not pd.isnull(df_07_data.iloc[int_07_result_num_indx, 1]):
                    if df_07_data.iloc[int_07_result_num_indx, 1] in [0, 1]:
                        pass
                    else:
                        int_07_result_num_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[6] + "中第" + str(int_07_result_num_indx + 1) +
                            "行第B列存在异常，异常原因：'结果判定'项内容不合规")
                else:
                    int_07_result_num_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[6] + "中第" + str(int_07_result_num_indx + 1) +
                        "行第B列存在异常，异常原因：'结果判定'项内容为空")
        else:
            int_07_result_num_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[6] + "中第1行存在异常，异常原因：未找到名为'结果判定'的项")
        # if int_07_result_num_stat:
            # logger.debug(
            #     "'结果判定'信息位于" + CHECK_SHEET_LIST[6] + "中第2到" + str(df_07_data.shape[0]) + "行第2列，校验无误")
        # 校验结果判定

        if int_06_test_type_stat:
            if set(df_07_data.iloc[0, 2:]) == set(df_06_data.iloc[1:, 0]):
                int_07_test_type_stat = 1
                # logger.debug(
                #     "'测试项目'信息位于" + CHECK_SHEET_LIST[6] + "中第1行第3到" + str(df_07_data.shape[1]) +
                #     "列，校验无误")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[6] + "中第1行第3到" + get_column_letter(df_07_data.shape[1]) +
                    "列存在异常，异常原因：'测试项目'项内容与" + CHECK_SHEET_LIST[5] + "中不相匹")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[5] + "中'测试项目'信息异常，无法完成'测试项目'校验")
        # 校验测试项目

        if int_07_test_result_stat and int_07_test_type_stat:
            for int_07_area_result_type_indx in range(1, df_07_data.shape[0]):
                int_07_row_result_type_stat = 0
                for int_07_area_result_type_col in range(2, df_07_data.shape[1]):
                    if pd.isnull(df_07_data.iloc[int_07_area_result_type_indx, int_07_area_result_type_col]):
                        pass
                    else:
                        int_07_row_result_type_stat = 1
                        if df_07_data.iloc[int_07_area_result_type_indx, int_07_area_result_type_col] == "X":
                            pass
                        else:
                            int_07_area_result_type_stat = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[6] + "中第" + str(int_07_area_result_type_indx + 1) + "行第" +
                                get_column_letter(int_07_area_result_type_col + 1) + "列存在异常，异常原因：内容无效，有效值为'X'")
                if not int_07_row_result_type_stat:
                    int_07_area_result_type_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[6] + "中第" + str(int_07_area_result_type_indx + 1) + "行第C到" +
                        get_column_letter(df_07_data.shape[1]) + "列存在异常，异常原因：未找到该行'测试结果-测试项目'信息")
        else:
            int_07_area_result_type_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[6] + "中'测试结果'、'测试项目'信息异常，无法完成'测试结果-测试项目'校验")
        # if int_07_area_result_type_stat:
            # logger.debug(
            #     "'测试结果-测试项目'位于" + CHECK_SHEET_LIST[6] + "中第2到" + str(df_07_data.shape[0]) + "行第3到" +
            #     str(df_05_data.shape[1]) + "列，校验无误")
        # 校验测试结果-测试项目

        if int_mod_test_stat:
            if int_07_test_result_stat and int_07_result_num_stat:
                dict_07_result_num = df_07_data.iloc[1:, 0:2].set_index(0).T.to_dict('list')
                for str_mod_test in list_mod_test:
                    if str_mod_test in dict_07_result_num.keys():
                        if dict_07_result_num[str_mod_test][0] == 1:
                            pass
                        else:
                            int_07_link_result_num_stat = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[6] + "中第A列存在异常，异常原因：'测试结果'列内容为" +
                                str_mod_test + "的项对应'结果判定'值不合规，应为1")
                    else:
                        int_07_link_result_num_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[6] + "中第A列存在异常，异常原因：'测试结果'列内无法找到" +
                            CHECK_SHEET_LIST[0] + "中内容为" + str_mod_test + "的'测试异常'项")
            else:
                int_07_link_result_num_stat = 0
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    CHECK_SHEET_LIST[6] + "中'测试结果'、'结果判定'信息异常，无法完成'测试结果-结果判定'校验")
        else:
            int_07_link_result_num_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[0] + "中'测试异常'信息异常，无法完成'测试结果-结果判定'校验")
        # if int_07_link_result_num_stat:
            # logger.debug(
            #     CHECK_SHEET_LIST[0] + "中'测试异常'信息皆包含于" + CHECK_SHEET_LIST[6] + "中第2到" +
            #     str(df_07_data.shape[0]) + "行第1到2列的'测试结果-结果判定'关系内，校验无误")
        # 校验测试结果-结果判定

        if int_07_test_result_stat and int_07_result_num_stat and int_07_test_type_stat \
                and int_07_area_result_type_stat and int_07_link_result_num_stat:
            for int_07_link_num_type_col in range(2, df_07_data.shape[1]):
                list_07_link_num_type = df_07_data[
                    int_07_link_num_type_col][df_07_data[int_07_link_num_type_col].isin(["X"])].index.tolist()
                list_07_col_link = df_07_data.iloc[list_07_link_num_type, 1].values.tolist()
                if list_07_col_link.count(0) == 1:
                    if list_07_col_link.count(1) >= 1:
                        pass
                    else:
                        int_07_link_num_type_stat = 0
                        msg_storage, dict_error = when_error(
                            msg_storage,
                            dict_error,
                            CHECK_SHEET_LIST[6] + "中第B列存在异常，异常原因：未找到" +
                            str(df_07_data.iloc[0, int_07_link_num_type_col]) +
                            "列对应'测试结果'和'结果判定'为异常的项")
                else:
                    int_07_link_num_type_stat = 0
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        CHECK_SHEET_LIST[6] + "中第B列存在异常，异常原因：" +
                        str(df_07_data.iloc[0, int_07_link_num_type_col]) +
                        "列对应'测试结果'和'结果判定'为正常的项应有且仅有一个")
        else:
            int_07_link_num_type_stat = 0
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                CHECK_SHEET_LIST[6] + "中存在信息异常，无法完成'结果判定-测试项目'校验")
        # if int_07_link_result_num_stat:
            # logger.debug(
            #     "各'测试项目'对应的'测试结果'判定同时包含异常和正常结果，且异常结果的数量≥1，正常结果的数量为1")
        # 校验结果判定-测试项目
    else:
        int_07_test_result_stat = 0
        int_07_result_num_stat = 0
        int_07_area_result_type_stat = 0
        int_07_link_result_num_stat = 0
        int_07_link_num_type_stat = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[6])
    stat_07 = \
        int_07_test_result_stat and int_07_result_num_stat and int_07_test_type_stat and \
        int_07_area_result_type_stat and int_07_link_result_num_stat and int_07_link_num_type_stat
    # 校验07TestResult-Test#############################################################################################

    stat_mod_n_frame = 0
    stat_mod_n_fm2sy = 1
    stat_mod_n_fm2op = 1
    stat_mod_n_ip2sy = 1
    stat_mod_n_ip2op = 1
    if CHECK_SHEET_LIST[8] in list_sheet_name:
        msg_storage = when_info(msg_storage, "----------开始校验" + CHECK_SHEET_LIST[8] + "表----------")
        df_mod_n_data = odd_xl_data[CHECK_SHEET_LIST[8]]
        ws_mod_n = wb_xl[CHECK_SHEET_LIST[8]]

        if stat_mod:
            if df_mod_n_data.shape == df_mod_data.shape:
                if df_mod_n_data.iloc[list_fs_loc[0], list_fs_loc[1]] == "FS" and \
                        df_mod_n_data.iloc[list_fo_loc[0], list_fo_loc[1]] == "FO" and \
                        df_mod_n_data.iloc[list_is_loc[0], list_is_loc[1]] == "IS" and \
                        df_mod_n_data.iloc[list_io_loc[0], list_io_loc[1]] == "IO":
                    stat_mod_n_frame = 1
                else:
                    msg_storage, dict_error = when_error(
                        msg_storage,
                        dict_error,
                        "Model_N表中'FS'、'FO'、'IS'、'IO'位置与Model_Y中不符，无法进行校核")
            else:
                msg_storage, dict_error = when_error(
                    msg_storage,
                    dict_error,
                    "Model_N表与Model_Y表长宽不符，无法进行校核")
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "Model_Y表信息异常，无法校核Model_N表")
        # 校验Model_N框架格式

        if stat_mod_n_frame:
            for top_row in range(0, 3):
                for top_col in range(0, df_mod_n_data.shape[1]):
                    ws_mod_n.cell(row=top_row + 1, column=top_col + 1).value = \
                        ws_mod.cell(row=top_row + 1, column=top_col + 1).value
            for side_row in range(0, df_mod_n_data.shape[0]):
                for side_col in range(0, 3):
                    ws_mod_n.cell(row=side_row + 1, column=side_col + 1).value = \
                        ws_mod.cell(row=side_row + 1, column=side_col + 1).value
            # 框架值复制
            for fm2sy_row in range(3, list_io_loc[0]):
                for fm2sy_col in range(3, list_io_loc[1]):
                    if df_mod_data.iloc[fm2sy_row, fm2sy_col] == "X":
                        if df_mod_n_data.iloc[fm2sy_row, fm2sy_col] in [0, 1, "X"]:
                            pass
                        else:
                            stat_mod_n_fm2sy = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[8] + "中第" + str(fm2sy_row + 1) + "行第" +
                                get_column_letter(fm2sy_col + 1) + "列存在异常，异常原因：内容无效，有效值为0、1、'X'")
                            ws_mod_n.cell(row=fm2sy_row + 1, column=fm2sy_col + 1).fill = FILL_ERROR
                    else:
                        if df_mod_n_data.iloc[fm2sy_row, fm2sy_col] == "" or \
                                df_mod_n_data.iloc[fm2sy_row, fm2sy_col] is None or \
                                pd.isnull(df_mod_n_data.iloc[fm2sy_row, fm2sy_col]):
                            pass
                        else:
                            stat_mod_n_fm2sy = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[8] + "中第" + str(fm2sy_row + 1) + "行第" +
                                get_column_letter(fm2sy_col + 1) + "列存在异常，异常原因：内容与Model_Y中不符，应为空")
                            ws_mod_n.cell(row=fm2sy_row + 1, column=fm2sy_col + 1).fill = FILL_WARNING
                            ws_mod_n.cell(row=fm2sy_row + 1, column=fm2sy_col + 1).value = ""
            # fm2sy关联校验
            for fm2op_row in range(3, list_io_loc[0]):
                for fm2op_col in range(list_io_loc[1] + 1, df_mod_n_data.shape[1]):
                    if df_mod_data.iloc[fm2op_row, fm2op_col] == "X":
                        if df_mod_n_data.iloc[fm2op_row, fm2op_col] in [0, 1, "X"]:
                            pass
                        else:
                            stat_mod_n_fm2op = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[8] + "中第" + str(fm2op_row + 1) + "行第" +
                                get_column_letter(fm2op_col + 1) + "列存在异常，异常原因：内容无效，有效值为0、1、'X'")
                            ws_mod_n.cell(row=fm2op_row + 1, column=fm2op_col + 1).fill = FILL_ERROR
                    else:
                        if df_mod_n_data.iloc[fm2op_row, fm2op_col] == "" or \
                                df_mod_n_data.iloc[fm2op_row, fm2op_col] is None or \
                                pd.isnull(df_mod_n_data.iloc[fm2op_row, fm2op_col]):
                            pass
                        else:
                            stat_mod_n_fm2op = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[8] + "中第" + str(fm2op_row + 1) + "行第" +
                                get_column_letter(fm2op_col + 1) + "列存在异常，异常原因：内容与Model_Y中不符，应为空")
                            ws_mod_n.cell(row=fm2op_row + 1, column=fm2op_col + 1).fill = FILL_WARNING
                            ws_mod_n.cell(row=fm2op_row + 1, column=fm2op_col + 1).value = ""
            # fm2op关联校验
            for ip2sy_row in range(list_io_loc[0] + 1, df_mod_n_data.shape[0]):
                for ip2sy_col in range(3, list_io_loc[1]):
                    if df_mod_data.iloc[ip2sy_row, ip2sy_col] == "X":
                        if df_mod_n_data.iloc[ip2sy_row, ip2sy_col] in [0, 1, "X"]:
                            pass
                        else:
                            stat_mod_n_ip2sy = 0
                            msg_storage, dict_error = when_error(
                                msg_storage,
                                dict_error,
                                CHECK_SHEET_LIST[8] + "中第" + str(ip2sy_row + 1) + "行第" +
                                get_column_letter(ip2sy_col + 1) + "列存在异常，异常原因：内容无效，有效值为0、1、'X'")
                            ws_mod_n.cell(row=ip2sy_row + 1, column=ip2sy_col + 1).fill = FILL_ERROR
                    else:
                        if df_mod_n_data.iloc[ip2sy_row, ip2sy_col] == "" or \
                                df_mod_n_data.iloc[ip2sy_row, ip2sy_col] is None or \
                                pd.isnull(df_mod_n_data.iloc[ip2sy_row, ip2sy_col]):
                            pass
                        else:
                            stat_mod_n_ip2sy = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[8] + "中第" + str(ip2sy_row + 1) + "行第" +
                                get_column_letter(ip2sy_col + 1) + "列存在异常，异常原因：内容与Model_Y中不符，应为空")
                            ws_mod_n.cell(row=ip2sy_row + 1, column=ip2sy_col + 1).fill = FILL_WARNING
                            ws_mod_n.cell(row=ip2sy_row + 1, column=ip2sy_col + 1).value = ""
            # ip2sy关联校验
            for ip2op_row in range(list_io_loc[0] + 1, df_mod_n_data.shape[0]):
                for ip2op_col in range(list_io_loc[1] + 1, df_mod_n_data.shape[1]):
                    if df_mod_data.iloc[ip2op_row, ip2op_col] == "X":
                        if df_mod_n_data.iloc[ip2op_row, 0] == df_mod_n_data.iloc[0, ip2op_col] and \
                                df_mod_n_data.iloc[ip2op_row, 1] == df_mod_n_data.iloc[1, ip2op_col]:
                            if df_mod_n_data.iloc[ip2op_row, ip2op_col] == 1:
                                pass
                            else:
                                stat_mod_n_ip2op = 0
                                msg_storage, dict_warning = when_warning(
                                    msg_storage,
                                    dict_warning,
                                    CHECK_SHEET_LIST[8] + "中第" + str(ip2op_row + 1) + "行第" +
                                    get_column_letter(ip2op_col + 1) + "列存在异常，异常原因：内容无效，有效值为1")
                                ws_mod_n.cell(row=ip2op_row + 1, column=ip2op_col + 1).fill = FILL_WARNING
                                ws_mod_n.cell(row=ip2op_row + 1, column=ip2op_col + 1).value = 1
                        else:
                            if df_mod_n_data.iloc[ip2op_row, ip2op_col] in [0, 1, "X"]:
                                pass
                            else:
                                stat_mod_n_ip2op = 0
                                msg_storage, dict_warning = when_warning(
                                    msg_storage,
                                    dict_warning,
                                    CHECK_SHEET_LIST[8] + "中第" + str(ip2op_row + 1) + "行第" +
                                    get_column_letter(ip2op_col + 1) + "列存在异常，异常原因：内容无效，有效值为0、1、'X'")
                                ws_mod_n.cell(row=ip2op_row + 1, column=ip2op_col + 1).fill = FILL_ERROR
                    else:
                        if df_mod_n_data.iloc[ip2op_row, ip2op_col] == "" or \
                                df_mod_n_data.iloc[ip2op_row, ip2op_col] is None or \
                                pd.isnull(df_mod_n_data.iloc[ip2op_row, ip2op_col]):
                            pass
                        else:
                            stat_mod_n_ip2op = 0
                            msg_storage, dict_warning = when_warning(
                                msg_storage,
                                dict_warning,
                                CHECK_SHEET_LIST[8] + "中第" + str(ip2op_row + 1) + "行第" +
                                get_column_letter(ip2op_col + 1) + "列存在异常，异常原因：内容与Model_Y中不符，应为空")
                            ws_mod_n.cell(row=ip2op_row + 1, column=ip2op_col + 1).fill = FILL_WARNING
                            ws_mod_n.cell(row=ip2op_row + 1, column=ip2op_col + 1).value = ""
            # ip2op关联校验
        else:
            stat_mod_n_fm2sy = 0
            stat_mod_n_fm2op = 0
            stat_mod_n_ip2sy = 0
            stat_mod_n_ip2op = 0
    else:
        stat_mod_n_fm2sy = 0
        stat_mod_n_fm2op = 0
        stat_mod_n_ip2sy = 0
        stat_mod_n_ip2op = 0
        msg_storage, dict_warning = when_warning(msg_storage, dict_warning, "未找到" + CHECK_SHEET_LIST[8])
    stat_mod_n = stat_mod_n_frame and stat_mod_n_fm2sy and stat_mod_n_fm2op and stat_mod_n_ip2sy and stat_mod_n_ip2op
    # 校验Model_N#######################################################################################################

    if CHECK_SHEET_LIST[7] in list_sheet_name:
        if stat_mod and stat_mod_n:
            wb_xl.remove(wb_xl[CHECK_SHEET_LIST[7]])
            msg_storage = when_info(msg_storage, "----------开始生成" + CHECK_SHEET_LIST[7] + "表----------")
            ws_mod_sum = wb_xl.copy_worksheet(ws_mod_n)
            ws_mod_sum.title = CHECK_SHEET_LIST[7]

            for fm2sy_row in range(3, list_io_loc[0]):
                for fm2sy_col in range(3, list_io_loc[1]):
                    if df_mod_n_data.iloc[fm2sy_row, fm2sy_col] in [0, 1, "X"]:
                        ws_mod_sum.cell(row=fm2sy_row + 1, column=fm2sy_col + 1).value = \
                            "[X," + str(df_mod_n_data.iloc[fm2sy_row, fm2sy_col]) + "]"
                    else:
                        ws_mod_sum.cell(row=fm2sy_row + 1, column=fm2sy_col + 1).value = "[,]"
            # fm2sy关联合并
            for fm2op_row in range(3, list_io_loc[0]):
                for fm2op_col in range(list_io_loc[1] + 1, df_mod_n_data.shape[1]):
                    if df_mod_n_data.iloc[fm2op_row, fm2op_col] in [0, 1, "X"]:
                        ws_mod_sum.cell(row=fm2op_row + 1, column=fm2op_col + 1).value = \
                            "[X," + str(df_mod_n_data.iloc[fm2op_row, fm2op_col]) + "]"
                    else:
                        ws_mod_sum.cell(row=fm2op_row + 1, column=fm2op_col + 1).value = "[,]"
            # fm2op关联合并
            for ip2sy_row in range(list_io_loc[0] + 1, df_mod_n_data.shape[0]):
                for ip2sy_col in range(3, list_io_loc[1]):
                    if df_mod_n_data.iloc[ip2sy_row, ip2sy_col] in [0, 1, "X"]:
                        ws_mod_sum.cell(row=ip2sy_row + 1, column=ip2sy_col + 1).value = \
                            "[X," + str(df_mod_n_data.iloc[ip2sy_row, ip2sy_col]) + "]"
                    else:
                        ws_mod_sum.cell(row=ip2sy_row + 1, column=ip2sy_col + 1).value = "[,]"
            # ip2sy关联合并
            for ip2op_row in range(list_io_loc[0] + 1, df_mod_n_data.shape[0]):
                for ip2op_col in range(list_io_loc[1] + 1, df_mod_n_data.shape[1]):
                    if df_mod_n_data.iloc[ip2op_row, ip2op_col] in [0, 1, "X"]:
                        ws_mod_sum.cell(row=ip2op_row + 1, column=ip2op_col + 1).value = \
                            "[X," + str(df_mod_n_data.iloc[ip2op_row, ip2op_col]) + "]"
                    else:
                        ws_mod_sum.cell(row=ip2op_row + 1, column=ip2op_col + 1).value = "[,]"
            # ip2op关联合并
        else:
            msg_storage, dict_error = when_error(
                msg_storage,
                dict_error,
                "Model_Y、Model_N表信息异常，无法校核Model表")
    # 校验Model(sum)####################################################################################################

    if stat_01 and stat_mod and stat_03 and stat_04 and stat_05 and stat_06 and stat_07 and stat_mod_n:
        msg_storage = when_info(
            msg_storage,
            "====================已完成EXCEL文件：" + str_xl_name + "的模型校核，未发现异常====================\n")
        return 1, msg_storage, dict_warning, dict_error, wb_xl
    else:
        return 0, msg_storage, dict_warning, dict_error, wb_xl


# if __name__ == '__main__':
#     str_filePath = r"C:\Users\shaoweigu\Desktop\Diagnose\checkMod\data\2020_01_04_17_42"
#     str_savePath = r"C:\Users\shaoweigu\Desktop\Diagnose\checkMod\result\2020_01_04_17_42"
#     int_checkLV = 1
#
#     list_xlName = select_excel(str_filePath, int_checkLV)
#     for str_xlName in list_xlName:
#         int_pointLoc = re.search(r".x", str_xlName).span()[0]
#         logID = logger.add(str_savePath + "/" + str_xlName[0:int_pointLoc]+".log", level="INFO")
#         # logID = logger.add(str_savePath + "/" + str_xlName[0:int_pointLoc] + ".log")
#         odd_xlData = pd.read_excel(str_filePath + "/" + str_xlName, sheet_name=None, header=None)
#         renameStat = check_sheet(odd_xlData, str_xlName, logID)
#         if renameStat:
#             os.rename(
#                 str_savePath + "/" + str_xlName[0:int_pointLoc] + ".log",
#                 str_savePath + "/Passed-" + str_xlName[0:int_pointLoc] + ".log")
#         else:
#             os.rename(
#                 str_savePath + "/" + str_xlName[0:int_pointLoc] + ".log",
#                 str_savePath + "/Error-" + str_xlName[0:int_pointLoc] + ".log")
